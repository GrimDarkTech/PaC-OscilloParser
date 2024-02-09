import openpyxl

class ExcelManager:

    @staticmethod
    def open_book(bookName: str) -> openpyxl.Workbook:
        book = openpyxl.open(f"Excels/{bookName}.xlsx")
        return book

    @staticmethod
    def create_book(bookName: str) -> openpyxl.Workbook:
        book = openpyxl.Workbook()
        book.save(f"Excels/{bookName}.xlsx")
        return book

    @staticmethod
    def write_to_excel(bookName: str, title: str , values, column_num = 1):
        # column_num - Номер столбца, в который будем записывать значения
        # values - Массив значений
        # file_path - Имя файла

        # Открываем файл Excel
        wb = ExcelManager.open_book(bookName)

        # Выбираем активный лист
        sheet = wb.active
        #Write title
        sheet.cell(row=1, column=column_num).value = title
        # Проходим по значениям и записываем их во второй столбец
        for row_num, value in enumerate(values, start=2):
            # Записываем значение в указанную ячейку
            sheet.cell(row=row_num, column=column_num).value = value

        # Сохраняем изменения в файл
        wb.save(f"Excels/{bookName}.xlsx")
        wb.close()
